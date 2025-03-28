from database import get_db_connection
import os
from openpyxl import load_workbook
import asyncio


EXCEL_PATH = "src/promocodes.xlsx"
CAMPAIGN = "Water pro"

async def import_promocodes_from_excel():
    try:
        conn = await get_db_connection()

        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active

        inserted = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header row
            promocode, points = row
            if not promocode or not isinstance(points, int):
                continue
            try:
                await conn.execute(
                    "INSERT INTO promocodes (promocode, campaign, points) VALUES ($1, $2, $3)",
                    promocode.strip(), CAMPAIGN, points
                )
                inserted += 1
            except Exception as e:
                print(f"❌ Skipped [{promocode}]: {e}")

        await conn.close()
        print(f"✅ Imported {inserted} promocodes for campaign '{CAMPAIGN}'")
    except Exception as ex:
        print(f"error {ex}")


async def main():
    await import_promocodes_from_excel()

asyncio.run(main())
